#!/usr/bin/env python3
"""Clean and summarize ecommerce comments from text, CSV, TSV, JSON, or XLSX."""

from __future__ import annotations

import argparse
import csv
import json
import re
from collections import Counter
from pathlib import Path
from typing import Iterable


TOPICS = {
    "口味口感": ("好吃", "难吃", "口味", "口感", "甜", "酸", "咸", "腻", "香", "脆", "糯"),
    "配料健康": ("配料", "成分", "添加", "糖", "脂肪", "热量", "蛋白", "健康", "过敏"),
    "包装规格": ("包装", "破损", "漏", "瓶", "袋", "盒", "大份", "小份", "分量", "规格"),
    "价格促销": ("价格", "贵", "便宜", "优惠", "活动", "性价比", "券"),
    "物流售后": ("物流", "快递", "发货", "客服", "售后", "退款", "退货"),
    "复购推荐": ("回购", "复购", "推荐", "安利", "下次", "囤", "再买"),
    "场景人群": ("早餐", "办公室", "通勤", "夜宵", "健身", "孩子", "老人", "送礼", "聚会"),
}


def normalize(value: object) -> str:
    text = re.sub(r"\s+", " ", str(value or "")).strip()
    return text


def redact(text: str) -> str:
    text = re.sub(r"(?<!\d)1[3-9]\d{9}(?!\d)", "[手机号已隐藏]", text)
    text = re.sub(r"@[\w\u4e00-\u9fff.-]+", "@[用户名已隐藏]", text)
    return text


def choose_column(fieldnames: list[str], requested: str | None) -> str:
    if requested:
        if requested not in fieldnames:
            raise ValueError(f"找不到评论列 {requested!r}")
        return requested
    candidates = ("comment", "content", "text", "评论", "评论内容", "内容")
    for candidate in candidates:
        if candidate in fieldnames:
            return candidate
    if len(fieldnames) == 1:
        return fieldnames[0]
    raise ValueError("无法自动识别评论列，请使用 --text-column 指定")


def read_delimited(path: Path, delimiter: str, column: str | None) -> list[str]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle, delimiter=delimiter)
        if not reader.fieldnames:
            return []
        selected = choose_column(reader.fieldnames, column)
        return [normalize(row.get(selected)) for row in reader]


def read_json(path: Path, column: str | None) -> list[str]:
    data = json.loads(path.read_text(encoding="utf-8-sig"))
    items = data if isinstance(data, list) else data.get("comments", [])
    output = []
    for item in items:
        if isinstance(item, str):
            output.append(normalize(item))
        elif isinstance(item, dict):
            key = column or next((k for k in ("comment", "content", "text", "评论", "内容") if k in item), None)
            if key:
                output.append(normalize(item.get(key)))
    return output


def read_xlsx(path: Path, column: str | None) -> list[str]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError("读取XLSX需要安装 openpyxl，或先导出为CSV") from exc
    sheet = load_workbook(path, read_only=True, data_only=True).active
    rows = sheet.iter_rows(values_only=True)
    headers = [normalize(value) for value in next(rows, ())]
    selected = choose_column(headers, column)
    index = headers.index(selected)
    return [normalize(row[index] if index < len(row) else "") for row in rows]


def read_comments(path: Path, column: str | None) -> list[str]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return read_delimited(path, ",", column)
    if suffix == ".tsv":
        return read_delimited(path, "\t", column)
    if suffix == ".json":
        return read_json(path, column)
    if suffix == ".xlsx":
        return read_xlsx(path, column)
    if suffix in {".txt", ".md"}:
        return [normalize(line) for line in path.read_text(encoding="utf-8-sig").splitlines()]
    raise ValueError("支持 .txt、.md、.csv、.tsv、.json 和 .xlsx")


def meaningful(text: str) -> bool:
    compact = re.sub(r"[^\w\u4e00-\u9fff]", "", text)
    return len(compact) >= 2


def topic_counts(comments: Iterable[str]) -> tuple[Counter[str], dict[str, list[str]]]:
    counts: Counter[str] = Counter()
    examples: dict[str, list[str]] = {topic: [] for topic in TOPICS}
    for comment in comments:
        for topic, keywords in TOPICS.items():
            if any(keyword in comment for keyword in keywords):
                counts[topic] += 1
                if len(examples[topic]) < 3:
                    examples[topic].append(redact(comment[:120]))
    return counts, examples


def render(raw_count: int, comments: list[str]) -> str:
    counts, examples = topic_counts(comments)
    lines = [
        "# 评论初步分析", "", "## 样本概况", "",
        f"- 原始条数：{raw_count}", f"- 清洗去重后：{len(comments)}",
        f"- 删除/重复：{raw_count - len(comments)}", "",
        "> 本结果为关键词初筛，必须结合上下文进行语义复核，不能直接代表整体消费者观点。", "",
        "## 主题命中", "", "| 主题 | 命中评论数 | 样本占比 |", "|---|---:|---:|",
    ]
    for topic, count in counts.most_common():
        share = count / len(comments) * 100 if comments else 0
        lines.append(f"| {topic} | {count} | {share:.1f}% |")
    if not counts:
        lines.append("| 暂无关键词命中 | 0 | 0.0% |")
    lines.extend(["", "## 匿名化代表原话（待人工复核）", ""])
    for topic, _ in counts.most_common():
        lines.append(f"### {topic}")
        lines.append("")
        lines.extend(f"- “{text}”" for text in examples[topic])
        lines.append("")
    return "\n".join(lines).rstrip() + "\n"


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("input", type=Path)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--text-column", help="CSV/TSV/XLSX/JSON中的评论列名")
    args = parser.parse_args()
    try:
        raw = read_comments(args.input, args.text_column)
    except (ValueError, OSError, json.JSONDecodeError) as exc:
        raise SystemExit(str(exc)) from exc
    cleaned = list(dict.fromkeys(text for text in raw if meaningful(text)))
    args.output.write_text(render(len(raw), cleaned), encoding="utf-8")


if __name__ == "__main__":
    main()
